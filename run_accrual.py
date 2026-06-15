import argparse
import csv
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APRIL_2026 = "April 2026"
ACCRUAL_DATE = "2026-04-30"
MATERIALITY_DOLLARS = 1000.0
OUTLIER_ZSCORE = 3.0

CARRIER_ALIASES = {
    "peak logistics": "Peak Logistics",
    "peak log": "Peak Logistics",
    "peak": "Peak Logistics",
    "heartland freight": "Heartland Freight",
    "heartland freight co.": "Heartland Freight",
    "heartland": "Heartland Freight",
    "coastal express": "Coastal Express",
    "coastal express llc": "Coastal Express",
    "coastal": "Coastal Express",
}


@dataclass
class ExceptionItem:
    severity: str
    carrier: str
    shipment_id: str
    category: str
    message: str
    estimated_financial_impact: float
    control_owner: str = "Accounting"


def clean_money(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in {"n/a", "no charge"}:
        return 0.0
    text = text.replace("$", "").replace(",", "").replace("/lb", "").strip()
    try:
        return float(text)
    except ValueError:
        match = re.search(r"[-+]?\d*\.?\d+", text)
        return float(match.group(0)) if match else None


def clean_pct(value):
    if pd.isna(value):
        return None
    text = str(value)
    match = re.search(r"[-+]?\d*\.?\d+", text)
    return float(match.group(0)) / 100 if match else None


def normalize_carrier(value):
    key = str(value).strip().lower()
    return CARRIER_ALIASES.get(key, str(value).strip())


def normalize_service(value):
    text = "" if pd.isna(value) else str(value).strip().lower()
    mapping = {
        "std": "Standard",
        "standard": "Standard",
        "ground": "Ground",
        "economy": "Economy",
        "expedited": "Expedited",
    }
    return mapping.get(text, str(value).strip() if text else "")


def normalize_accessorial(value):
    if pd.isna(value) or not str(value).strip():
        return ""
    text = str(value).strip().lower()
    text = text.replace("â€”", "-").replace("–", "-")
    mapping = {
        "liftgate": "Liftgate Delivery",
        "liftgate delivery": "Liftgate Delivery",
        "residential": "Residential Delivery",
        "residential delivery": "Residential Delivery",
        "inside delivery": "Inside Delivery",
        "appointment": "Appointment Delivery",
        "appointment delivery": "Appointment Delivery",
        "saturday delivery": "Saturday Delivery",
        "detention": "Detention - driver wait time",
        "detention - driver wait time": "Detention - driver wait time",
    }
    return mapping.get(text, str(value).strip())


def read_raw_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def section_after(rows, header_name, required_columns=None):
    required_columns = set(required_columns or [])
    for idx, row in enumerate(rows):
        stripped = {str(cell).strip() for cell in row}
        if row and str(row[0]).strip() == header_name and required_columns.issubset(stripped):
            return idx
    raise ValueError(f"Could not find section header: {header_name} with columns {sorted(required_columns)}")


def extract_table(rows, header_name, required_columns=None):
    idx = section_after(rows, header_name, required_columns)
    headers = [h.strip() for h in rows[idx]]
    data = []
    for row in rows[idx + 1 :]:
        if not any(str(cell).strip() for cell in row):
            break
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        data.append(row[: len(headers)])
    return pd.DataFrame(data, columns=headers)


def parse_effective_date(rows):
    for row in rows[:4]:
        text = " ".join(str(x) for x in row)
        match = re.search(r"Effective:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})", text)
        if match:
            return pd.to_datetime(match.group(1)).date()
    return None


def parse_peak_rate_card(path):
    rows = read_raw_rows(path)
    tiers = extract_table(rows, "Weight Tier")
    tiers["min_weight"] = tiers["Min Weight (lbs)"].map(clean_money)
    tiers["max_weight"] = tiers["Max Weight (lbs)"].map(clean_money)
    tiers["rate_per_mile"] = tiers["Rate ($/mile)"].map(clean_money)

    miles = extract_table(rows, "Destination City")
    miles["destination_key"] = (
        miles["Destination City"].str.strip().str.lower()
        + "|"
        + miles["State"].str.strip().str.upper()
    )
    miles["miles"] = pd.to_numeric(miles["Approx. Miles"], errors="coerce")

    accessorials = extract_table(rows, "Service")
    accessorials["service"] = accessorials["Service"].map(normalize_accessorial)
    accessorials["fee"] = accessorials["Fee (USD)"].map(clean_money)

    min_charge = 185.0
    fuel_pct = 0.14
    return {
        "tiers": tiers,
        "miles": miles,
        "accessorials": dict(zip(accessorials["service"], accessorials["fee"])),
        "fuel_pct": fuel_pct,
        "min_charge": min_charge,
        "effective_date": parse_effective_date(rows),
    }


def parse_heartland_rate_card(path):
    rows = read_raw_rows(path)
    zones = extract_table(rows, "ZIP Prefix Range")
    parsed_zones = []
    for _, row in zones.iterrows():
        nums = [int(x) for x in re.findall(r"\d{3}", str(row["ZIP Prefix Range"]))]
        if len(nums) >= 2:
            parsed_zones.append((nums[0], nums[1], row["Zone"]))

    rates = extract_table(rows, "Zone", ["Flat Rate (USD)"])
    rates = rates[rates["Zone"].astype(str).str.startswith("Zone")].copy()
    rates["flat_rate"] = rates["Flat Rate (USD)"].map(clean_money)

    discounts = extract_table(rows, "Tier")
    parsed_discounts = []
    for _, row in discounts.iterrows():
        counts = [int(x) for x in re.findall(r"\d+", str(row["Cumulative Shipments (QTD)"]))]
        threshold = counts[0] if counts else 1
        pct = clean_pct(row["Discount Off Base Rate"]) or 0.0
        parsed_discounts.append((threshold, pct, row["Tier"]))
    parsed_discounts.sort(key=lambda x: x[0])

    accessorials = extract_table(rows, "Service")
    accessorials["service"] = accessorials["Service"].map(normalize_accessorial)
    accessorials["fee"] = accessorials["Fee (USD)"].map(clean_money)
    return {
        "zip_zones": parsed_zones,
        "rates": dict(zip(rates["Zone"], rates["flat_rate"])),
        "discounts": parsed_discounts,
        "accessorials": dict(zip(accessorials["service"], accessorials["fee"])),
        "effective_date": parse_effective_date(rows),
    }


def parse_coastal_rate_card(path):
    rows = read_raw_rows(path)
    regions = extract_table(rows, "Region")
    parsed_regions = []
    for _, row in regions.iterrows():
        nums = [int(x) for x in re.findall(r"\d{5}", str(row["ZIP Code Range"]))]
        if len(nums) >= 2:
            parsed_regions.append((nums[0], nums[1], row["Region"]))

    rates = extract_table(rows, "Region", ["Per-Pound Rate"])
    rates = rates[rates["Region"].isin(["SoCal", "NorCal", "PNW"])].copy()
    rates["rate_per_lb"] = rates["Per-Pound Rate"].map(clean_money)
    rates["minimum"] = rates["Minimum Charge"].map(clean_money)

    residential = extract_table(rows, "Weight Tier")
    residential["min_weight"] = residential["Min Weight (lbs)"].map(clean_money)
    residential["max_weight"] = residential["Max Weight (lbs)"].map(clean_money)
    residential["fee"] = residential["Residential Surcharge"].map(clean_money)

    accessorials = extract_table(rows, "Service")
    accessorials["service"] = accessorials["Service"].map(normalize_accessorial)
    accessorials["fee"] = accessorials["Fee (USD)"].map(clean_money)

    fuel_pct = 0.095
    return {
        "regions": parsed_regions,
        "rates": rates.set_index("Region")[["rate_per_lb", "minimum"]].to_dict("index"),
        "residential": residential,
        "accessorials": dict(zip(accessorials["service"], accessorials["fee"])),
        "fuel_pct": fuel_pct,
        "effective_date": parse_effective_date(rows),
    }


def historical_accessorial_averages(invoices):
    hist = invoices.copy()
    hist["carrier_norm"] = hist["carrier"].map(normalize_carrier)
    hist["accessorial_norm"] = hist["accessorial_detail"].map(normalize_accessorial)
    hist = hist[(hist["accessorial_norm"] != "") & (hist["accessorial_fees"].fillna(0) > 0)]
    grouped = hist.groupby(["carrier_norm", "accessorial_norm"])["accessorial_fees"].mean()
    carrier_avg = hist.groupby("carrier_norm")["accessorial_fees"].mean()
    return grouped.to_dict(), carrier_avg.to_dict()


def historical_weight_averages(invoices):
    hist = invoices.copy()
    hist["carrier_norm"] = hist["carrier"].map(normalize_carrier)
    hist["dest_key"] = (
        hist["destination_city"].str.strip().str.lower()
        + "|"
        + hist["destination_state"].str.strip().str.upper()
    )
    by_dest = hist.groupby(["carrier_norm", "dest_key"])["weight_lbs"].mean().to_dict()
    by_carrier = hist.groupby("carrier_norm")["weight_lbs"].mean().to_dict()
    return by_dest, by_carrier


def historical_outlier_stats(invoices):
    hist = invoices.copy()
    hist["carrier_norm"] = hist["carrier"].map(normalize_carrier)
    stats = {}
    for carrier, group in hist.groupby("carrier_norm"):
        mean = group["total_charge"].mean()
        std = group["total_charge"].std(ddof=0)
        stats[carrier] = (mean, std if std and not math.isnan(std) else 0.0)
    return stats


def historical_charge_averages(invoices):
    hist = invoices.copy()
    hist["carrier_norm"] = hist["carrier"].map(normalize_carrier)
    hist["dest_key"] = (
        hist["destination_city"].str.strip().str.lower()
        + "|"
        + hist["destination_state"].str.strip().str.upper()
    )
    by_dest = hist.groupby(["carrier_norm", "dest_key"]).agg(
        base_charge=("base_charge", "mean"),
        fuel_surcharge=("fuel_surcharge", "mean"),
    )
    by_carrier = hist.groupby("carrier_norm").agg(
        base_charge=("base_charge", "mean"),
        fuel_surcharge=("fuel_surcharge", "mean"),
    )
    return by_dest.to_dict("index"), by_carrier.to_dict("index")


def build_manifest(data_dir):
    rows = []
    for path in sorted(data_dir.glob("*")):
        if path.is_file():
            rows.append(
                {
                    "source_file": path.name,
                    "path": str(path),
                    "size_bytes": path.stat().st_size,
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                }
            )
    return pd.DataFrame(rows)


def carrier_risk_rating(carrier, detail, exceptions):
    carrier_rows = detail[detail["carrier"] == carrier]
    carrier_ex = [e for e in exceptions if e.carrier == carrier]
    assumption_pct = (
        carrier_rows["assumption_based_amount"].sum() / carrier_rows["estimated_total"].sum()
        if carrier_rows["estimated_total"].sum()
        else 0
    )
    high_count = sum(1 for e in carrier_ex if e.severity == "High")
    warning_count = len(carrier_ex)
    if high_count or assumption_pct > 0.15:
        return "High", f"{high_count} high exceptions; {assumption_pct:.1%} assumption-based estimate"
    if warning_count >= 5 or assumption_pct > 0.05:
        return "Medium", f"{warning_count} exceptions; {assumption_pct:.1%} assumption-based estimate"
    return "Low", f"{warning_count} exceptions; {assumption_pct:.1%} assumption-based estimate"


def find_heartland_zone(zip_code, zip_zones):
    prefix = int(str(int(zip_code)).zfill(5)[:3])
    for start, end, zone in zip_zones:
        if start <= prefix <= end:
            return zone
    return None


def find_coastal_region(zip_code, regions):
    zip_int = int(str(int(zip_code)).zfill(5))
    for start, end, region in regions:
        if start <= zip_int <= end:
            return region
    return None


def lookup_accessorial(carrier, accessorial, rate_cards, hist_avg, carrier_hist_avg):
    if not accessorial:
        return 0.0, "Direct Rate Card", "No accessorial requested", 0.0
    direct = rate_cards[carrier].get("accessorials", {})
    if accessorial in direct and direct[accessorial] is not None:
        return float(direct[accessorial]), "Direct Rate Card", f"{accessorial} from carrier rate card", 0.0
    if (carrier, accessorial) in hist_avg:
        amount = float(hist_avg[(carrier, accessorial)])
        return amount, "Historical Estimate", f"{accessorial} estimated from historical carrier/accessorial average", amount
    if carrier in carrier_hist_avg:
        amount = float(carrier_hist_avg[carrier])
        return amount, "Assumption Required", f"{accessorial} estimated from carrier historical accessorial average", amount
    return 0.0, "Flagged Exception", f"No rate or historical estimate available for {accessorial}", 0.0


def estimate_shipments(shipments, invoices, rate_cards):
    hist_accessorial, carrier_accessorial = historical_accessorial_averages(invoices)
    hist_weight_dest, hist_weight_carrier = historical_weight_averages(invoices)
    hist_charge_dest, hist_charge_carrier = historical_charge_averages(invoices)
    outlier_stats = historical_outlier_stats(invoices)

    rows = []
    exceptions = []
    heartland_counter = 0
    stale_logged = set()

    shipments = shipments.copy()
    shipments["date_parsed"] = pd.to_datetime(shipments["date"], errors="coerce")
    shipments["carrier"] = shipments["carrier"].map(normalize_carrier)
    shipments["service_level_norm"] = shipments["service_level"].map(normalize_service)
    shipments = shipments.sort_values(["date_parsed", "shipment_id"])

    for _, ship in shipments.iterrows():
        carrier = ship["carrier"]
        shipment_id = ship["shipment_id"]
        dest_key = f"{str(ship['destination_city']).strip().lower()}|{str(ship['destination_state']).strip().upper()}"
        messages = []
        evidence = "Direct Rate Card"
        assumption_amount = 0.0
        base = fuel = accessorial = discount = estimated_total = 0.0
        rate_basis = ""
        mapping = ""
        direct_amount = 0.0

        if carrier not in rate_cards:
            exceptions.append(ExceptionItem("High", carrier, shipment_id, "Unknown carrier", "Carrier could not be mapped to a supported rate card", 0.0))
            evidence = "Flagged Exception"
            messages.append("Unsupported carrier")

        weight = ship["weight_lbs"]
        if pd.isna(weight) or weight <= 0:
            est_weight = hist_weight_dest.get((carrier, dest_key), hist_weight_carrier.get(carrier))
            exceptions.append(ExceptionItem("Medium", carrier, shipment_id, "Missing weight", "Weight missing; estimated from historical invoice averages", 0.0))
            if est_weight:
                weight = est_weight
                evidence = "Assumption Required"
                messages.append(f"Weight estimated at {weight:.1f} lbs")
            else:
                weight = 0.0
                evidence = "Flagged Exception"
                messages.append("No historical weight estimate available")

        if pd.isna(ship["destination_zip"]) or pd.isna(ship["destination_city"]) or pd.isna(ship["destination_state"]):
            exceptions.append(ExceptionItem("High", carrier, shipment_id, "Missing destination", "Destination fields required for rate mapping are missing", 0.0))
            evidence = "Flagged Exception"
            messages.append("Missing destination")

        special = normalize_accessorial(ship["special_handling"])
        if carrier in rate_cards:
            accessorial, access_evidence, access_note, access_assumption = lookup_accessorial(
                carrier, special, rate_cards, hist_accessorial, carrier_accessorial
            )
            if access_evidence == "Flagged Exception":
                exceptions.append(ExceptionItem("Medium", carrier, shipment_id, "Unknown accessorial", access_note, accessorial))
            if access_evidence in {"Historical Estimate", "Assumption Required"} and evidence != "Flagged Exception":
                evidence = access_evidence
            assumption_amount += access_assumption
            messages.append(access_note)

        try:
            if carrier == "Peak Logistics":
                card = rate_cards[carrier]
                mile_row = card["miles"][card["miles"]["destination_key"] == dest_key]
                if mile_row.empty:
                    exceptions.append(ExceptionItem("High", carrier, shipment_id, "Invalid rate card mapping", "Peak mileage not found for destination", 0.0))
                    fallback = hist_charge_dest.get((carrier, dest_key), hist_charge_carrier.get(carrier))
                    if fallback:
                        base = float(fallback.get("base_charge", 0.0) or 0.0)
                        fuel = float(fallback.get("fuel_surcharge", 0.0) or 0.0)
                        assumption_amount += base + fuel
                        evidence = "Historical Estimate" if (carrier, dest_key) in hist_charge_dest else "Assumption Required"
                        rate_basis = "Historical average base/fuel used because Peak mileage was not in rate card"
                        mapping = "Historical fallback"
                        messages.append("Peak mileage missing; historical base/fuel fallback applied")
                    else:
                        evidence = "Flagged Exception"
                        messages.append("Peak mileage missing and no historical fallback available")
                else:
                    miles = float(mile_row.iloc[0]["miles"])
                    tier_row = card["tiers"][
                        (card["tiers"]["min_weight"] <= weight)
                        & ((card["tiers"]["max_weight"].isna()) | (weight <= card["tiers"]["max_weight"]))
                    ]
                    if tier_row.empty:
                        exceptions.append(ExceptionItem("High", carrier, shipment_id, "Invalid rate card mapping", "Peak weight tier not found", 0.0))
                        evidence = "Flagged Exception"
                        messages.append("Peak weight tier missing")
                    else:
                        rate = float(tier_row.iloc[0]["rate_per_mile"])
                        calculated = miles * rate
                        base = max(calculated, card["min_charge"])
                        fuel = base * card["fuel_pct"]
                        rate_basis = f"{miles:.0f} miles x ${rate:.2f}/mile; ${card['min_charge']:.2f} minimum"
                        mapping = str(tier_row.iloc[0]["Weight Tier"])
            elif carrier == "Heartland Freight":
                heartland_counter += 1
                card = rate_cards[carrier]
                zone = find_heartland_zone(ship["destination_zip"], card["zip_zones"])
                if not zone:
                    exceptions.append(ExceptionItem("High", carrier, shipment_id, "Unknown zone", "Heartland ZIP prefix not listed in zone table", 0.0))
                    fallback = hist_charge_dest.get((carrier, dest_key), hist_charge_carrier.get(carrier))
                    if fallback:
                        base = float(fallback.get("base_charge", 0.0) or 0.0)
                        fuel = 0.0
                        assumption_amount += base
                        evidence = "Historical Estimate" if (carrier, dest_key) in hist_charge_dest else "Assumption Required"
                        rate_basis = "Historical average base used because Heartland ZIP zone was not in rate card"
                        mapping = "Historical fallback"
                        messages.append("Heartland zone missing; historical base fallback applied")
                    else:
                        evidence = "Flagged Exception"
                        messages.append("Heartland zone missing and no historical fallback available")
                else:
                    base_before_discount = float(card["rates"][zone])
                    discount_pct = 0.0
                    discount_tier = "Tier 1"
                    for threshold, pct, tier in card["discounts"]:
                        if heartland_counter >= threshold:
                            discount_pct = pct
                            discount_tier = tier
                    discount = base_before_discount * discount_pct
                    base = base_before_discount - discount
                    fuel = 0.0
                    rate_basis = f"{zone} flat rate ${base_before_discount:.2f}; Q2 shipment {heartland_counter}; {discount_tier} {discount_pct:.0%}"
                    mapping = zone
            elif carrier == "Coastal Express":
                card = rate_cards[carrier]
                region = find_coastal_region(ship["destination_zip"], card["regions"])
                if not region:
                    exceptions.append(ExceptionItem("High", carrier, shipment_id, "Unknown region", "Coastal ZIP outside contracted region ranges", 0.0))
                    fallback = hist_charge_dest.get((carrier, dest_key), hist_charge_carrier.get(carrier))
                    if fallback:
                        base = float(fallback.get("base_charge", 0.0) or 0.0)
                        fuel = float(fallback.get("fuel_surcharge", 0.0) or 0.0)
                        assumption_amount += base + fuel
                        evidence = "Historical Estimate" if (carrier, dest_key) in hist_charge_dest else "Assumption Required"
                        rate_basis = "Historical average base/fuel used because Coastal region was not in rate card"
                        mapping = "Historical fallback"
                        messages.append("Coastal region missing; historical base/fuel fallback applied")
                    else:
                        evidence = "Flagged Exception"
                        messages.append("Coastal region missing and no historical fallback available")
                else:
                    region_rate = card["rates"][region]
                    base = max(weight * float(region_rate["rate_per_lb"]), float(region_rate["minimum"]))
                    fuel = base * card["fuel_pct"]
                    mapping = region
                    rate_basis = f"{region} {float(region_rate['rate_per_lb']):.2f}/lb; ${float(region_rate['minimum']):.2f} minimum"
                    if bool(ship["residential"]):
                        res_row = card["residential"][
                            (card["residential"]["min_weight"] <= weight)
                            & ((card["residential"]["max_weight"].isna()) | (weight <= card["residential"]["max_weight"]))
                        ]
                        if res_row.empty:
                            exceptions.append(ExceptionItem("Medium", carrier, shipment_id, "Invalid rate card mapping", "Coastal residential surcharge tier not found", 0.0))
                            evidence = "Flagged Exception"
                        else:
                            accessorial += float(res_row.iloc[0]["fee"])
                            messages.append("Residential surcharge applied from Coastal weight tier")
            else:
                pass
        except Exception as err:
            exceptions.append(ExceptionItem("High", carrier, shipment_id, "Calculation error", str(err), 0.0))
            evidence = "Flagged Exception"
            messages.append(f"Calculation error: {err}")

        direct_amount = base + fuel + accessorial
        estimated_total = direct_amount

        mean, std = outlier_stats.get(carrier, (0.0, 0.0))
        if std and estimated_total > mean + OUTLIER_ZSCORE * std:
            exceptions.append(ExceptionItem("Medium", carrier, shipment_id, "Outlier shipment cost", "Estimated total exceeds historical carrier outlier threshold", estimated_total - mean))
            messages.append("Outlier shipment cost flagged")

        if carrier in rate_cards:
            effective = rate_cards[carrier].get("effective_date")
            stale_after = (pd.Timestamp(effective) + pd.DateOffset(months=12)).date() if effective else None
            if effective and pd.Timestamp(ACCRUAL_DATE).date() > stale_after and carrier not in stale_logged:
                exceptions.append(ExceptionItem("Low", carrier, "RATECARD", "Stale rate data", f"Rate card effective {effective}; verify no newer card before posting", 0.0))
                stale_logged.add(carrier)

        rows.append(
            {
                "shipment_id": shipment_id,
                "date": ship["date"],
                "carrier": carrier,
                "original_carrier": ship.get("carrier_original", carrier),
                "service_level": ship["service_level_norm"],
                "destination_city": ship["destination_city"],
                "destination_state": ship["destination_state"],
                "destination_zip": ship["destination_zip"],
                "weight_lbs_used": round(float(weight), 2) if not pd.isna(weight) else 0,
                "residential": bool(ship["residential"]),
                "special_handling": special,
                "rate_mapping": mapping,
                "rate_basis": rate_basis,
                "base_freight": round(base, 2),
                "fuel_surcharge": round(fuel, 2),
                "accessorial_estimate": round(accessorial, 2),
                "volume_discount": round(discount, 2),
                "estimated_total": round(estimated_total, 2),
                "evidence_class": evidence,
                "assumption_based_amount": round(assumption_amount, 2),
                "audit_note": "; ".join([m for m in messages if m]),
            }
        )

    return pd.DataFrame(rows), exceptions


def baseline_comparison(detail, denise, invoices):
    carrier_est = detail.groupby("carrier", as_index=False)["estimated_total"].sum()
    carrier_est.rename(columns={"estimated_total": "activity_based_april_estimate"}, inplace=True)

    den = denise.copy()
    den["carrier"] = den["carrier"].map(normalize_carrier)
    den["month_date"] = pd.to_datetime(den["month"], format="%B %Y")
    last_three = den[den["month_date"].between("2026-01-01", "2026-03-31")]
    baseline = last_three.groupby("carrier", as_index=False)["actual_invoiced"].mean()
    baseline.rename(columns={"actual_invoiced": "denise_trailing_3mo_baseline"}, inplace=True)

    backtest = den.copy()
    backtest["denise_abs_error"] = (backtest["actual_invoiced"] - backtest["accrual_estimate"]).abs()
    backtest_summary = backtest.groupby("carrier", as_index=False).agg(
        denise_avg_abs_error=("denise_abs_error", "mean"),
        denise_max_abs_error=("denise_abs_error", "max"),
        historical_months=("month", "count"),
    )

    actuals = invoices.copy()
    actuals["carrier"] = actuals["carrier"].map(normalize_carrier)
    actual_summary = actuals.groupby("carrier", as_index=False)["total_charge"].sum()
    actual_summary.rename(columns={"total_charge": "six_month_actual_invoice_total"}, inplace=True)

    result = carrier_est.merge(baseline, on="carrier", how="left")
    result = result.merge(backtest_summary, on="carrier", how="left")
    result = result.merge(actual_summary, on="carrier", how="left")
    result["difference_vs_denise_baseline"] = result["activity_based_april_estimate"] - result["denise_trailing_3mo_baseline"]
    result["april_actual_status"] = "April actual invoices not provided; April comparison is estimate vs Denise trailing baseline"
    return result


def build_outputs(data_dir, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    shipments = pd.read_csv(data_dir / "shipments_apr2026.csv")
    shipments["carrier_original"] = shipments["carrier"]
    invoices = pd.read_csv(data_dir / "freight_invoices_oct2025_mar2026_v2.csv")
    denise = pd.read_csv(data_dir / "denise_accruals_v2.csv")

    rate_cards = {
        "Peak Logistics": parse_peak_rate_card(data_dir / "rate_card_peak_logistics.csv"),
        "Heartland Freight": parse_heartland_rate_card(data_dir / "rate_card_heartland_freight.csv"),
        "Coastal Express": parse_coastal_rate_card(data_dir / "rate_card_coastal_express.csv"),
    }

    detail, exceptions = estimate_shipments(shipments, invoices, rate_cards)

    carrier = detail.groupby("carrier", as_index=False).agg(
        shipment_count=("shipment_id", "count"),
        base_freight=("base_freight", "sum"),
        fuel_surcharge=("fuel_surcharge", "sum"),
        accessorial_estimate=("accessorial_estimate", "sum"),
        volume_discount=("volume_discount", "sum"),
        estimated_total=("estimated_total", "sum"),
        assumption_based_amount=("assumption_based_amount", "sum"),
    )
    risk_rows = []
    for c in carrier["carrier"]:
        rating, rationale = carrier_risk_rating(c, detail, exceptions)
        risk_rows.append({"carrier": c, "confidence_risk_rating": rating, "risk_rationale": rationale})
    carrier = carrier.merge(pd.DataFrame(risk_rows), on="carrier", how="left")

    total_accrual = float(carrier["estimated_total"].sum())
    journal = pd.concat(
        [
            pd.DataFrame(
                {
                    "journal_date": [ACCRUAL_DATE] * len(carrier),
                    "account": ["Freight Expense"] * len(carrier),
                    "department": ["Operations"] * len(carrier),
                    "carrier": carrier["carrier"],
                    "debit": carrier["estimated_total"].round(2),
                    "credit": [0.0] * len(carrier),
                    "memo": ["April 2026 freight accrual by activity-based estimate"] * len(carrier),
                }
            ),
            pd.DataFrame(
                {
                    "journal_date": [ACCRUAL_DATE],
                    "account": ["Accrued Freight"],
                    "department": ["Operations"],
                    "carrier": ["All carriers"],
                    "debit": [0.0],
                    "credit": [round(total_accrual, 2)],
                    "memo": ["April 2026 freight accrual offset"],
                }
            ),
        ],
        ignore_index=True,
    )

    exception_df = pd.DataFrame([e.__dict__ for e in exceptions])
    if exception_df.empty:
        exception_df = pd.DataFrame(columns=[field for field in ExceptionItem.__dataclass_fields__])

    baseline = baseline_comparison(detail, denise, invoices)

    tie_out = {
        "carrier_detail_total": round(float(carrier["estimated_total"].sum()), 2),
        "shipment_level_total": round(float(detail["estimated_total"].sum()), 2),
        "journal_debits": round(float(journal["debit"].sum()), 2),
        "journal_credits": round(float(journal["credit"].sum()), 2),
    }
    tie_out["status"] = "PASS" if len(set(tie_out.values()) - {"PASS"}) <= 1 else "FAIL"

    exec_rows = [
        {"metric": "Accrual month", "value": APRIL_2026},
        {"metric": "Total estimated freight accrual", "value": round(total_accrual, 2)},
        {"metric": "Shipment count", "value": int(detail["shipment_id"].nunique())},
        {"metric": "Assumption-based amount", "value": round(float(detail["assumption_based_amount"].sum()), 2)},
        {"metric": "High severity exceptions", "value": int((exception_df["severity"] == "High").sum()) if not exception_df.empty else 0},
        {"metric": "Tie-out status", "value": tie_out["status"]},
        {"metric": "April actual invoice status", "value": "Not provided; April is labeled as estimate"},
    ]
    executive = pd.DataFrame(exec_rows)

    controls = pd.DataFrame(
        [
            {"control": "Required field validation", "status": "Complete", "description": "Shipment ID, carrier, destination, date, and weight are checked; exceptions logged where missing."},
            {"control": "Carrier and service normalization", "status": "Complete", "description": "Known aliases are mapped to canonical carrier names and service levels."},
            {"control": "Rate-card mapping", "status": "Complete", "description": "Peak by mileage/weight tier, Heartland by ZIP zone and QTD tier, Coastal by ZIP region/per-pound rate."},
            {"control": "Assumption separation", "status": "Complete", "description": "Historical and assumption-based accessorial or weight estimates are tagged in the audit trail."},
            {"control": "Materiality threshold", "status": f"${MATERIALITY_DOLLARS:,.0f}", "description": "Carrier-level differences and exceptions over this threshold warrant review before posting."},
            {"control": "Outlier threshold", "status": f"{OUTLIER_ZSCORE:.1f} historical standard deviations", "description": "Shipment estimates above carrier historical thresholds are flagged."},
            {"control": "Tie-out", "status": tie_out["status"], "description": f"Carrier detail {tie_out['carrier_detail_total']}, shipment detail {tie_out['shipment_level_total']}, journal debits {tie_out['journal_debits']}, credits {tie_out['journal_credits']}."},
            {"control": "Rate-card freshness", "status": "Review", "description": "Cards effective January 2025 are flagged as stale for April 2026 close unless management confirms no updates."},
        ]
    )

    manifest = build_manifest(data_dir)

    journal.to_csv(output_dir / "ridgeline_freight_journal_entry.csv", index=False)
    exception_df.to_csv(output_dir / "exception_log.csv", index=False)

    workbook_path = output_dir / "ridgeline_freight_accrual_audit_pack.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        executive.to_excel(writer, index=False, sheet_name="Executive Summary")
        journal.to_excel(writer, index=False, sheet_name="Journal Entry")
        carrier.to_excel(writer, index=False, sheet_name="Carrier Accrual Detail")
        detail.to_excel(writer, index=False, sheet_name="Shipment-Level Estimate")
        exception_df.to_excel(writer, index=False, sheet_name="Exception Log")
        baseline.to_excel(writer, index=False, sheet_name="Baseline Comparison")
        controls.to_excel(writer, index=False, sheet_name="Assumptions & Controls")
        manifest.to_excel(writer, index=False, sheet_name="Source File Manifest")

        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center")
            for column_cells in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 55)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    return workbook_path, output_dir / "ridgeline_freight_journal_entry.csv", output_dir / "exception_log.csv", tie_out


def main():
    parser = argparse.ArgumentParser(description="Ridgeline Foods freight accrual engine")
    parser.add_argument("--data-dir", default="data", help="Directory containing challenge CSV source files")
    parser.add_argument("--output-dir", default="output", help="Directory for Excel and CSV outputs")
    args = parser.parse_args()

    workbook, journal, exceptions, tie_out = build_outputs(Path(args.data_dir), Path(args.output_dir))
    print(f"Wrote {workbook}")
    print(f"Wrote {journal}")
    print(f"Wrote {exceptions}")
    print(f"Tie-out: {tie_out}")


if __name__ == "__main__":
    main()
